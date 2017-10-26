__author__ = ("Matthias Rost <mrost AT inet.tu-berlin.de>, "
              "Alexander Elvers <aelvers AT inet.tu-berlin.de>")

__all__ = [
    "PersonalPlan",
    "get_empty_plan",
    "get_target_plan",
    "compute_workload",
    "compute_workload_first_week",
    "compute_workload_second_week",
    "compute_happiness",
    "count_available_tutors",
    "count_available_rooms",
    "export_plan_to_xlsx",
    "import_plan_from_xlsx",
    "check_plan",
    "get_plan_paths",
    "save_plan_paths",
    "get_new_plan_folder",
]

import contextlib
import datetime
import itertools
import pathlib
import yaml
from functools import reduce
from typing import Dict, Set, List, Iterable, Tuple, Optional, Any, cast, Iterator

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet import Worksheet

from .data import Data
from .rooms import Room
from .tutor import Tutor
from ..util import converter, settings
from ..util.settings import DAYS, hours_real, TASKS, TUTORIUM, UEBUNG_MAR, UEBUNG_TEL, KONTROLLE


# type aliases
# task type -> day index -> hour -> count
PlanDict = Dict[str, Dict[int, Dict[int, int]]]
# tutor name -> task type -> day index -> hour -> room name or empty string
PersonalPlanDict = Dict[str, Dict[str, Dict[int, Dict[int, str]]]]


class PersonalPlan:
    """
    A personal plan contains all information about the TASKS of all tutors,
    all rooms and all days.

    At a time slot, a tutor can have at most one task. If the room is a
    tutorial room, it can also have at most one task. The tutor has to be
    available in regard to their availability preferences. The room has to be
    booked.
    """

    plan_by_tutor: Dict[Tutor, Dict[datetime.date, Dict[int, Room]]]
    plan_by_room: Dict[Room, Dict[datetime.date, Dict[int, Set[Tutor]]]]

    def __init__(self) -> None:
        """
        Create empty personal plan.
        """
        # plans by tutor and by room for faster access (O(1))
        self.plan_by_tutor = {}
        self.plan_by_room = {}

    def add_task(self, tutor: Tutor, date: datetime.date, time: int, room: Room) -> None:
        """
        Add task.
        """
        tutor_has_task = time in self.plan_by_tutor.get(tutor, {}).get(date, {})
        if tutor_has_task:
            raise ValueError("f{tutor} has already a task at {date} {time}")
        tutor_is_available = (tutor.availability.get(date, {}).get(time - time % 2) or 0) > 0
        if not tutor_is_available:
            raise ValueError("f{tutor} is unavailable at {date} {time}")
        room_is_booked = room.is_booked(date, time - time % 2)
        if not room_is_booked:
            raise ValueError("f{room} is not booked at {date} {time}")
        if room.type == "tutorial":
            # no double task for tutorial rooms
            room_has_task = time in self.plan_by_room.get(room, {}).get(date, {})
            if room_has_task:
                raise ValueError("f{room} has already a task at {date} {time}")

        self.plan_by_tutor.setdefault(tutor, {}).setdefault(date, {})[time] = room
        self.plan_by_room.setdefault(room, {}).setdefault(date, {}).setdefault(time, set()).add(tutor)

    def remove_task(self, tutor: Tutor, date: datetime.date, time: int, room: Optional[Room] = None) -> Room:
        """
        Remove task.

        The task has to exist. If the room is not given, it is not checked.

        Returns the room.
        """
        current_room = self.plan_by_tutor.get(tutor, {}).get(date, {}).get(time)
        if current_room is None:
            raise ValueError("f{tutor} has no task at {date} {time}")
        elif room is not None and current_room != room:
            raise ValueError("ftask of {tutor} at {date} {time} is not in {room}")
        del self.plan_by_tutor[tutor][date][time]
        self.plan_by_room[current_room][date][time].remove(tutor)
        return current_room

    def get_personal_plan(self) -> PersonalPlanDict:
        """
        Get the plan in the personal plan format that is used for export.

        The personal plan format is:
        tutor -> task type -> day index -> hour -> room or empty string
        """
        result: PersonalPlanDict = {}
        for tutor in Data().tutor_by_name.values():
            tutor_plan: Dict[str, Dict[int, Dict[int, str]]] = {}
            result[tutor.last_name] = tutor_plan
            for slot_type in [KONTROLLE, TUTORIUM, UEBUNG_MAR, UEBUNG_TEL]:
                tutor_plan[slot_type] = {day: {hour: "" for hour in range(10, 18)} for day in range(1, 11)}
            for date, day_plan in self.plan_by_tutor.get(tutor, {}).items():
                day_index = converter.date_to_day_index(date)
                for hour, room in day_plan.items():
                    tutor_plan[type_map[room.type]][day_index][hour] = room.name
        return result

    @classmethod
    def create_from_personal_plan(cls, personal_plan: PersonalPlanDict) -> "PersonalPlan":
        """
        Create a plan from the personal plan format that is used for export.

        The personal plan format is:
        tutor -> task type -> day index -> hour -> room or empty string
        """
        plan = cls()
        for tutor_name, tutor_plan in personal_plan.items():
            for type_plan in tutor_plan.values():
                for day_index, day_plan in type_plan.items():
                    date = converter.day_index_to_date(day_index)
                    for hour, room_name in day_plan.items():
                        if room_name:
                            plan.add_task(Data().tutor_by_name[tutor_name], date, hour, Data().room_by_name[room_name])
        return plan


def get_empty_plan() -> PlanDict:
    """
    Get an empty plan.
    """
    plan: PlanDict = {}
    for task in TASKS:
        plan[task] = {}
        for day in DAYS:
            plan[task][day] = {hour: 0 for hour in hours_real(day)}
    return plan


def get_target_plan() -> PlanDict:
    """
    Get the target plan.

    This loads the target plan from xlsx.
    """
    return import_plan_from_xlsx()


def print_single_time_row(table, day, name):
    list_of_values = [str(table[day][hour]) for hour in hours_real(day)]
    header = [name, ","]
    header.extend(list_of_values)
    print(",".join(header))


def print_master_plan(master_plan, ordered=True):
    element_names = [TUTORIUM, "tutorienRaeumeAnzahlen", UEBUNG_TEL, "poolAnzahlenTEL", UEBUNG_MAR, "poolAnzahlenMAR", KONTROLLE, "tutorenavailability"]

    if ordered:
        for day in range(1,11,1):
            print("\n\n\n\n\n\n\tTAG" + str(day) + "\n\n")
            for element_name in element_names:
                print_single_time_row(master_plan[element_name], day, element_name)
    else:
        for day in range(1,11,1):
            print("\n\n\n\n\n\n\tTAG" + str(day) + "\n\n")
            for element_name in sorted(master_plan.keys()):
                print_single_time_row(master_plan[element_name], day, element_name)




def compute_workload(plan: PlanDict) -> int:
    sum = 0
    for task in TASKS:
        for day in DAYS:
            for hour in hours_real(day):
                if plan[task][day][hour]:
                    sum += 1
    return sum


def compute_workload_first_week(plan: PlanDict) -> int:
    sum = 0
    for task in TASKS:
        for day in range(1,6,1):
            for hour in hours_real(day):
                if plan[task][day][hour]:
                    sum += 1
    return sum


def compute_workload_second_week(plan: PlanDict) -> int:
    sum = 0
    for task in TASKS:
        for day in range(6,11,1):
            for hour in hours_real(day):
                if plan[task][day][hour]:
                    sum += 1
    return sum


def compute_happiness(plan: PlanDict, availability: Dict[int, Dict[int, int]]) -> float:
    sum = 0.0
    for task in TASKS:
        for day in DAYS:
            for hour in hours_real(day):
                if plan[task][day][hour]:
                    sum += availability[day][hour]

    if compute_workload(plan) > 0:
        return sum / compute_workload(plan)
    return 666


def count_available_tutors(tutors: Iterable[Tutor], day: datetime.date, time: int) -> int:
    count = 0
    for tutor in tutors:
        if (tutor.availability[day][time] or 0) > 0:
            count += 1
    return count


def count_available_rooms(rooms: Iterable[Room], room_type: str, day: datetime.date, time: int) -> Tuple[int, int]:
    if time - (time % 2) in settings.settings.forbidden_timeslots._get(day, [])():  # TODO: dirty fix
        return 0, 0
    count = 0
    capacity = 0
    for room in rooms:
        if room.is_booked(day, time - (time % 2)) and room.type == room_type:  # TODO: dirty fix
            count += 1
            capacity += room.capacity
    return count, capacity


# TODO
type_map = dict(tutorial=TUTORIUM, exercise=UEBUNG_TEL, exerciseMAR=UEBUNG_MAR, grading=KONTROLLE)
type_map_inverse = {v: k for k, v in type_map.items()}


# slot_types = [TUTORIUM, UEBUNG_MAR, UEBUNG_TEL, KONTROLLE]
slot_types = ["tutorial", "exercise", "exerciseMAR", "grading"]


def write_plan_to_worksheet(ws: Worksheet, plan: PlanDict) -> None:
    """
    Write plan to worksheet.
    """
    times: List[int] = settings.settings.times()
    times = reduce(list.__add__, [[t, t+1] for t in times])
    days: List[datetime.date] = settings.settings.days()
    forbidden_timeslots = settings.settings.forbidden_timeslots()

    tutors = Data().tutor_by_name.values()
    rooms_ = Data().room_by_name.values()

    empty_lines = 2

    # width and height of one block
    width = 1 + len(times)
    height = 1 + len(slot_types) + 5 + empty_lines

    # write blocks
    columns: List[List[Optional[Any]]]
    for day in days:
        day_index = converter.date_to_day_index(day)
        # header
        ws.append([day] + times)
        # body
        columns = [cast(List[Optional[Any]], slot_types) + [None, "avail. tutors", "tutorial rooms", "tutorial seats", "exercise seats", "exerciseMAR seats"]]
        for time in times:
            column = []
            avail_tutors = count_available_tutors(tutors, day, time)
            avail_tutorials = count_available_rooms(rooms_, "tutorial", day, time)
            avail_exercises = count_available_rooms(rooms_, "exercise", day, time)
            avail_exercises_mar = count_available_rooms(rooms_, "exerciseMAR", day, time)
            avail_grading = count_available_rooms(rooms_, "grading", day, time)

            for slot_type in slot_types:
                avail = count_available_rooms(rooms_, slot_type, day, time)[0]
                column.append(plan[type_map[slot_type]][day_index].get(time) if avail > 0 else None)
            column.extend([None, avail_tutors, avail_tutorials[0], avail_tutorials[1], avail_exercises[1], avail_exercises_mar[1]])
            # column.extend([
            #     0 if avail_tutorials[0] > 0 else None,
            #     0 if avail_exercises[0] > 0 else None,
            #     0 if avail_grading[0] > 0 else None,
            #     None, avail_tutors, avail_tutorials[0], avail_tutorials[1], avail_exercises[1]])
            columns.append(column)
        for rows in list(itertools.zip_longest(*columns)):
            ws.append(rows)
        # space
        for i in range(empty_lines):
            ws.append([])

    ws.column_dimensions["A"].width = 20


@contextlib.contextmanager
def export_plan_to_xlsx(path: Optional[str] = None) -> Iterator[Workbook]:
    """
    Write plan to xlsx.
    """
    if path is None:
        path = settings.settings.paths.planner()

    try:
        wb = load_workbook(path)
    except IOError:
        wb = Workbook()
        del wb["Sheet"]

    yield wb

    wb.save(path)


def import_plan_from_xlsx() -> PlanDict:
    """
    Read plan from xlsx.
    """
    path = settings.settings.paths.planner()
    times = reduce(list.__add__, [[t, t + 1] for t in settings.settings.times()])
    days = settings.settings.days()
    forbidden_timeslots = settings.settings.forbidden_timeslots()

    all_rooms = Data().room_by_name.values()

    empty_lines = 2

    # width and height of one block
    width = 1 + len(times)
    height = 1 + len(slot_types) + 1 + 5 + empty_lines

    wb = load_workbook(path)
    ws = wb["Target"]

    # data = list(map(lambda x: list(map(lambda y: y.value, x)), ws))
    data = [[x.value for x in row] for row in ws]

    # plan = {t: {} for t in slot_types}
    plan: PlanDict = {t: {} for t in [TUTORIUM, UEBUNG_MAR, UEBUNG_TEL, KONTROLLE]}  # TODO

    for day in days:
        day_index = converter.date_to_day_index(day)
        offset = (day_index - 1) * height
        # check header
        header = data[offset][:1+len(times)]
        expected_header = [datetime.datetime.combine(day, datetime.time())] + times
        if header != expected_header:
            raise ValueError(f"wrong header at {day}: expected {expected_header} but was {header}")
        for i, slot_type in enumerate(slot_types):
            line = data[offset+i+1][:1+len(times)]
            if line[0] != slot_type:
                raise ValueError(f"wrong slot type at {day}: expected {slot_type} but was {line[0]}")
            # TODO
            old_slot_type = slot_type
            slot_type = type_map[slot_type]
            day_plan: Dict[int, int] = {}
            plan[slot_type][day_index] = day_plan
            for time, value in zip(times, line[1:1+len(times)]):
                if value is None:
                    value = 0
                avail_rooms = count_available_rooms(all_rooms, old_slot_type, day, time)
                if value > 0 and avail_rooms[0] == 0:
                    print(f"\033[1;31mWARNING: {day} {time} {old_slot_type}:"
                          f" value set to {value} but no room available or forbidden timeslot\033[0m")
                    value = 0
                day_plan[time] = int(value)
        # # TODO
        # slot_type = UEBUNG_MAR
        # plan[slot_type][day_index] = day_plan = {}
        # for time in times:
        #     day_plan[time] = 0
    return plan


def check_plan(plan: PersonalPlanDict, rooms: Iterable[Room]) -> List[Tuple[datetime.date, int, str, str, str]]:
    """
    Check if plan is correct. A valid plan only uses booked rooms.

    Returns tuples of (day, time, room, tutor, message) if plan is not correct.
    """
    rooms_dict = {r.name: r for r in rooms}
    errors = []
    for tutor, tutor_plan in plan.items():
        for task, task_plan in tutor_plan.items():
            for day_index, day_plan in task_plan.items():
                day = converter.day_index_to_date(day_index)
                for time, room_name in day_plan.items():
                    if not room_name:
                        continue
                    if room_name not in rooms_dict:
                        errors.append((day, time, room_name, tutor, "room does not exist"))
                        continue
                    room = rooms_dict[room_name]
                    booked = room.is_booked(day, time - (time % 2))
                    if not booked:
                        errors.append((day, time, room_name, tutor, "room is not booked"))
    return errors


# get and set plan paths from config file

def get_plan_paths() -> Dict[str, Optional[pathlib.Path]]:
    """
    Get the plan paths (active and working) from plan_paths.yaml. If active
    or working does not exist or the file is missing, use None as value.
    """
    plans_folder = pathlib.Path(settings.settings.paths._get("plans", "plans")())
    plan_paths: Dict[str, Optional[pathlib.Path]] = {"active": None, "working": None, "parent": None}
    try:
        with open(plans_folder / settings.PLAN_PATHS_FILE) as f:
            folder_names = yaml.safe_load(f)
        plan_paths.update([(k, plans_folder / v) for k, v in folder_names.items() if v])
    except FileNotFoundError:
        pass
    if plan_paths["active"] is not None:
        try:
            plan_paths["parent"] = plans_folder / (plan_paths["active"] / "parent_plan").read_text().rstrip("\n")
        except FileNotFoundError:
            pass
    return plan_paths


def save_plan_paths(plan_paths: Dict[str, Optional[pathlib.Path]]) -> None:
    """
    Save plan paths (active and working) to plan_paths.yaml.
    """
    if "parent" in plan_paths:
        del plan_paths["parent"]
    plans_folder = pathlib.Path(settings.settings.paths._get("plans", "plans")())
    folder_names = {k: str(v.relative_to(plans_folder)) for k, v in plan_paths.items() if v}
    with open(plans_folder / settings.PLAN_PATHS_FILE, "w") as f:
        yaml.safe_dump(folder_names, f, default_flow_style=False)


def get_new_plan_folder(label: str) -> pathlib.Path:
    """
    Get the next plan folder. The name format is YYYY-MM-DD-N-label
    where YYYY-MM-DD is the current date, N is an incremental number and label
    is a user defined label.

    The incremental number is calculated from the last plan by date.

    The label should be `initial` for initial planning, `rolling` for
    rolling-wave planning and `manual-updates` for manual edits using the
    update-plan command.

    The folder has to be created outside.
    """
    date = datetime.date.today()
    n = 1

    plans_folder = pathlib.Path(settings.settings.paths._get("plans", "plans")())
    try:
        sub_folders = map(lambda x: x.parts[-1].split("-")[3:4], plans_folder.iterdir())
        n = 1 + max((int(x[0]) for x in sub_folders if x and x[0].isdigit()), default=0)
    except FileNotFoundError:
        pass  # plans folder is missing

    return plans_folder / f"{date}-{n}-{label}"
